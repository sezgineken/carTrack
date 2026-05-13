from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


def style_header_row(sheet):
    header_font = Font(bold=True, size=12)
    header_fill = PatternFill(fill_type='solid', fgColor='FFF2CC')
    header_alignment = Alignment(horizontal='center', vertical='center')
    for cell in sheet[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment


def style_id_column(sheet, column_index=1):
    id_fill = PatternFill(fill_type='solid', fgColor='FFF2F2F2')
    id_alignment = Alignment(horizontal='center', vertical='center')
    for row_idx in range(1, sheet.max_row + 1):
        id_cell = sheet.cell(row=row_idx, column=column_index)
        id_cell.alignment = id_alignment
        id_cell.fill = id_fill


def apply_thin_black_borders(sheet):
    thin_side = Side(style='thin', color='FF000000')
    thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
        for cell in row:
            cell.border = thin_border


def autosize_columns(sheet, min_width=10, max_width=80, padding=2):
    for col_idx in range(1, sheet.max_column + 1):
        max_length = 0
        for row_idx in range(1, sheet.max_row + 1):
            value = sheet.cell(row=row_idx, column=col_idx).value
            value_str = '' if value is None else str(value)
            if len(value_str) > max_length:
                max_length = len(value_str)
        adjusted = min(max(max_length + padding, min_width), max_width)
        sheet.column_dimensions[get_column_letter(col_idx)].width = adjusted
