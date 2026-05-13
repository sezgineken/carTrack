from django.http import HttpResponse
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from .admin_helpers import apply_thin_black_borders, autosize_columns, style_header_row, style_id_column


def build_reservations_excel_response(rows):
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = (
        f'attachment; filename="arac_rezervasyonlari_{timezone.localtime().strftime("%Y%m%d_%H%M%S")}.xlsx"'
    )
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Arac Rezervasyonlari"
    headers = [
        'ID', 'Araç', 'Kullanıcı', 'Başlangıç Zamanı', 'Bitiş Zamanı', 'Amaç',
        'Aktif', 'Durum', 'Kullanım', 'Alış KM', 'Bırakış KM', 'Teslim Formu', 'Oluşturulma',
    ]
    sheet.append(headers)
    style_header_row(sheet)
    for idx, row in enumerate(rows, start=1):
        sheet.append([idx, *row])
    sheet.freeze_panes = 'A2'
    sheet.auto_filter.ref = f"A1:{get_column_letter(sheet.max_column)}{sheet.max_row}"
    style_id_column(sheet)
    apply_thin_black_borders(sheet)
    autosize_columns(sheet)
    workbook.save(response)
    return response
