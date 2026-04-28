from urllib.parse import quote, urlencode

from django import forms
from django.contrib import admin, messages
from django.http import HttpResponse
from django.shortcuts import redirect
from django.template.response import TemplateResponse
from django.urls import path, reverse
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from .admin_helpers import apply_thin_black_borders, autosize_columns, style_header_row, style_id_column
from .models import Vehicle, VehicleHistory, VehicleHistoryDocument, VehicleNote


class VehicleAdminForm(forms.ModelForm):
    """Admin form for Vehicle."""

    show_in_pool_choice = forms.TypedChoiceField(
        label='Araç, havuzda listelensin mi?',
        choices=((True, 'Evet'), (False, 'Hayır')),
        coerce=lambda value: str(value).lower() == 'true',
        widget=forms.RadioSelect,
        required=True,
    )
    operational_status_choice = forms.ChoiceField(
        label='Araç Durumu',
        choices=Vehicle.OPERATIONAL_STATUS_CHOICES,
        widget=forms.RadioSelect,
        required=True,
    )
    ownership_choice = forms.ChoiceField(
        label='Araç Aidiyeti',
        choices=Vehicle.OWNERSHIP_CHOICES,
        widget=forms.Select(attrs={'class': 'vTextField', 'style': 'max-width: 260px;'}),
        required=True,
    )

    class Meta:
        model = Vehicle
        fields = '__all__'

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.fields['show_in_pool_choice'].initial = True
        self.fields['operational_status_choice'].initial = 'ACTIVE'
        self.fields['ownership_choice'].initial = 'PERSONAL'
        instance = getattr(self, 'instance', None)
        if instance and instance.pk:
            self.fields['show_in_pool_choice'].initial = bool(instance.show_in_pool)
            self.fields['operational_status_choice'].initial = instance.operational_status
            self.fields['ownership_choice'].initial = instance.ownership

    def save(self, commit=True):
        vehicle = super().save(commit=commit)
        vehicle.show_in_pool = self.cleaned_data.get('show_in_pool_choice', True)
        vehicle.operational_status = self.cleaned_data.get('operational_status_choice', 'ACTIVE')
        vehicle.ownership = self.cleaned_data.get('ownership_choice', 'PERSONAL')
        if commit:
            vehicle.save(update_fields=['show_in_pool', 'operational_status', 'ownership'])
        return vehicle


class VehicleAdmin(admin.ModelAdmin):
    """Vehicle Admin"""

    form = VehicleAdminForm
    change_form_template = 'admin/vehicles/vehicle/change_form.html'
    change_list_template = 'admin/vehicles/vehicle/change_list.html'
    list_display = [
        'plate',
        'brand_model',
        'history_user_display',
        'current_kilometer_display',
        'location_display',
    ]
    list_filter = ['is_active', 'location', 'operational_status', 'ownership']
    search_fields = ['plate', 'brand_model']
    fieldsets = (
        ('Temel Bilgiler', {'fields': ('plate', 'brand_model', 'photo', 'show_in_pool_choice', 'operational_status_choice')}),
        (
            'Kilometre Bilgisi',
            {
                'fields': ('initial_kilometer',),
                'description': 'Aracın sisteme eklenirkenki kilometresi. İlk kullanımda bu değer otomatik olarak alınacaktır.',
            },
        ),
        (
            'Aracın Lokasyonu',
            {
                'fields': ('location',),
                'description': 'Araç anasayfadaki ilgili lokasyon kartı altında bu seçime göre listelenir.',
            },
        ),
        (
            'Araç Aidiyeti',
            {
                'fields': ('ownership_choice',),
                'description': 'Şahsi/Kiralık sekmelerine göre admin araç listesinde ayrılır.',
            },
        ),
    )

    def status_display(self, obj):
        if obj.operational_status == 'PASSIVE':
            return "⛔ Kullanım Dışı"
        if obj.is_available():
            return "✅ Müsait"
        usage = obj.get_current_usage()
        if usage:
            return f"🔴 {usage.user.get_full_name()}"
        return "🔴 Kullanımda"

    status_display.short_description = 'Durum'

    def operational_status_display(self, obj):
        return obj.get_operational_status_display()

    operational_status_display.short_description = 'Araç Durumu'

    def ownership_display(self, obj):
        return obj.get_ownership_display()

    ownership_display.short_description = 'Aidiyet'

    def vehicle_location_display(self, obj):
        return obj.get_location_display()

    vehicle_location_display.short_description = 'Lokasyon'

    def location_display(self, obj):
        return obj.get_location_display()

    location_display.short_description = 'LOKASYON'

    def current_kilometer_display(self, obj):
        # Admin form edits update initial_kilometer, so list should read the same source.
        return obj.initial_kilometer

    current_kilometer_display.short_description = "KİLOMETRE"

    def history_user_display(self, obj):
        history = getattr(obj, 'history', None)
        if history and history.user_name:
            return history.user_name
        return '-'

    history_user_display.short_description = 'KULLANICI'

    def get_queryset(self, request):
        qs = super().get_queryset(request)
        return qs.select_related('history')

    def changelist_view(self, request, extra_context=None):
        active_ownership = request.GET.get('ownership__exact')
        if active_ownership not in {'PERSONAL', 'RENTAL'}:
            active_ownership = 'PERSONAL'
            mutable_get = request.GET.copy()
            mutable_get['ownership__exact'] = active_ownership
            request.GET = mutable_get
            request.META['QUERY_STRING'] = mutable_get.urlencode()

        base_params = request.GET.copy()
        if 'ownership__exact' in base_params:
            base_params.pop('ownership__exact')

        personal_params = base_params.copy()
        personal_params['ownership__exact'] = 'PERSONAL'
        rental_params = base_params.copy()
        rental_params['ownership__exact'] = 'RENTAL'
        tabs = [
            {'key': 'PERSONAL', 'label': 'Şahsi Araçlar', 'url': f'?{urlencode(personal_params, doseq=True)}'},
            {'key': 'RENTAL', 'label': 'Kiralık Araçlar', 'url': f'?{urlencode(rental_params, doseq=True)}'},
        ]

        extra_context = extra_context or {}
        extra_context['ownership_tabs'] = tabs
        extra_context['active_ownership_tab'] = active_ownership
        return super().changelist_view(request, extra_context=extra_context)

    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path('export/', self.admin_site.admin_view(self.export_vehicles_excel_view), name='vehicles_vehicle_export'),
            path('<path:object_id>/history/', self.admin_site.admin_view(self.vehicle_history_view), name='vehicles_vehicle_history'),
            path(
                '<path:object_id>/notes/download/',
                self.admin_site.admin_view(self.download_vehicle_notes_view),
                name='vehicles_vehicle_notes_download',
            ),
        ]
        return custom_urls + urls

    def export_vehicles_excel_view(self, request):
        # Export should respect active filters/search, but not ownership tabs.
        # This ensures location-based exports include both personal and rental vehicles.
        export_request = request
        export_params = request.GET.copy()
        export_params.pop('ownership__exact', None)
        export_request.GET = export_params

        changelist = self.get_changelist_instance(export_request)
        vehicles = (
            changelist.get_queryset(export_request)
            .select_related('history')
            .prefetch_related('history__documents')
        )
        timestamp = timezone.localtime().strftime('%Y%m%d_%H%M%S')
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="araclar_{timestamp}.xlsx"'
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Araclar"
        headers = [
            'ID', 'Plaka', 'Marka/Model', 'Lokasyon', 'Aidiyet', 'Araç Durumu', 'Aktif', 'Havuzda Görünür',
            'Başlangıç Kilometresi', 'Model Yılı', 'Renk', 'Araç Sahibi/Şahıs', 'Kullanıcı Şirket / Departman',
            'Kullanıcı', 'Muayene Tarihi', 'Muayene Kalan Gün', 'Sigorta Bitiş Tarihi', 'Sigorta Kalan Gün',
            'Kasko Bitiş Tarihi', 'Kasko Kalan Gün', 'Ruhsat Belgeleri', 'Sigorta/Kasko Belgeleri',
            'En Son Bakım Kilometresi', 'Lastik Ebatı', 'Son Kullanıcı', 'Tarihçe Güncelleme Zamanı',
        ]
        sheet.append(headers)
        style_header_row(sheet)

        for idx, vehicle in enumerate(vehicles, start=1):
            history = getattr(vehicle, 'history', None)
            registration_docs = ''
            insurance_casco_docs = ''
            if history:
                docs = list(history.documents.all())
                registration_docs = ' | '.join(
                    doc.original_filename or doc.file.name
                    for doc in docs
                    if doc.document_type == VehicleHistoryDocument.TYPE_REGISTRATION
                )
                insurance_casco_docs = ' | '.join(
                    doc.original_filename or doc.file.name
                    for doc in docs
                    if doc.document_type == VehicleHistoryDocument.TYPE_INSURANCE_CASCO
                )

            sheet.append([
                idx,
                vehicle.plate,
                vehicle.brand_model,
                vehicle.get_location_display(),
                vehicle.get_ownership_display(),
                vehicle.get_operational_status_display(),
                'Evet' if vehicle.is_active else 'Hayır',
                'Evet' if vehicle.show_in_pool else 'Hayır',
                vehicle.initial_kilometer,
                history.model_year if history else '',
                history.color if history else '',
                history.owner_person if history else '',
                history.user_company_department if history else '',
                history.user_name if history else '',
                history.inspection_date.strftime('%d.%m.%Y') if history and history.inspection_date else '',
                history.inspection_remaining_days if history and history.inspection_remaining_days is not None else '',
                history.insurance_end_date.strftime('%d.%m.%Y') if history and history.insurance_end_date else '',
                history.insurance_remaining_days if history and history.insurance_remaining_days is not None else '',
                history.casco_end_date.strftime('%d.%m.%Y') if history and history.casco_end_date else '',
                history.casco_remaining_days if history and history.casco_remaining_days is not None else '',
                registration_docs,
                insurance_casco_docs,
                history.last_maintenance_kilometer if history else '',
                history.tire_size if history else '',
                history.last_user_notes if history else '',
                timezone.localtime(history.updated_at).strftime('%d.%m.%Y %H:%M') if history and history.updated_at else '',
            ])

        sheet.freeze_panes = 'A2'
        sheet.auto_filter.ref = f"A1:{get_column_letter(sheet.max_column)}{sheet.max_row}"
        style_id_column(sheet)
        apply_thin_black_borders(sheet)
        autosize_columns(sheet)
        workbook.save(response)
        return response

    def vehicle_history_view(self, request, object_id):
        vehicle = self.get_object(request, object_id)
        if vehicle is None:
            return redirect('admin:vehicles_vehicle_changelist')

        history, _ = VehicleHistory.objects.get_or_create(vehicle=vehicle)

        class VehicleHistoryForm(forms.ModelForm):
            model_year = forms.ChoiceField(
                label='Model Yılı',
                required=False,
                choices=[('', '---------')] + [(str(y), str(y)) for y in range(2000, 2051)],
                widget=forms.Select(attrs={'class': 'vTextField', 'style': 'max-width: 220px;'}),
            )

            class Meta:
                model = VehicleHistory
                fields = [
                    'model_year',
                    'color',
                    'owner_person',
                    'user_company_department',
                    'user_name',
                    'inspection_date',
                    'insurance_end_date',
                    'casco_end_date',
                    'last_maintenance_kilometer',
                    'last_user_notes',
                    'tire_size',
                ]
                widgets = {
                    'color': forms.Select(attrs={'class': 'vTextField', 'style': 'max-width: 260px;'}),
                    'owner_person': forms.TextInput(attrs={'style': 'max-width: 520px; width: 100%;'}),
                    'user_company_department': forms.TextInput(attrs={'style': 'max-width: 520px; width: 100%;'}),
                    'user_name': forms.TextInput(attrs={'style': 'max-width: 520px; width: 100%;'}),
                    'inspection_date': forms.DateInput(attrs={'type': 'date', 'style': 'max-width: 220px;'}, format='%Y-%m-%d'),
                    'insurance_end_date': forms.DateInput(attrs={'type': 'date', 'style': 'max-width: 220px;'}, format='%Y-%m-%d'),
                    'casco_end_date': forms.DateInput(attrs={'type': 'date', 'style': 'max-width: 220px;'}, format='%Y-%m-%d'),
                    'last_maintenance_kilometer': forms.NumberInput(attrs={'min': 0, 'style': 'max-width: 220px;'}),
                    'last_user_notes': forms.Textarea(attrs={'rows': 4, 'style': 'max-width: 720px; width: 100%;'}),
                    'tire_size': forms.TextInput(attrs={'style': 'max-width: 320px; width: 100%;'}),
                }

            def clean_model_year(self):
                value = self.cleaned_data.get('model_year')
                if value in (None, '', '---------'):
                    return None
                try:
                    year = int(value)
                except (TypeError, ValueError):
                    raise forms.ValidationError('Geçerli bir yıl seçiniz.')
                if year < 2000 or year > 2050:
                    raise forms.ValidationError('Yıl 2000-2050 aralığında olmalıdır.')
                return year

        if request.method == 'POST' and request.POST.get('note_action') == 'add_note':
            note_text = (request.POST.get('vehicle_note_text') or '').strip()
            if not note_text:
                self.message_user(request, 'Boş not kaydedilemez.', level=messages.ERROR)
            else:
                VehicleNote.objects.create(
                    vehicle=vehicle,
                    note_text=note_text,
                    created_by=request.user if request.user.is_authenticated else None,
                )
                self.message_user(request, 'Araç notu kaydedildi.')
            return redirect(reverse('admin:vehicles_vehicle_history', args=[vehicle.pk]))

        if request.method == 'POST':
            form = VehicleHistoryForm(request.POST, request.FILES, instance=history)
            if form.is_valid():
                for field in ('inspection_date', 'insurance_end_date', 'casco_end_date'):
                    posted = request.POST.get(field)
                    if posted in (None, '') and getattr(history, field) is not None:
                        setattr(form.instance, field, getattr(history, field))
                form.save()

                delete_registration_ids = request.POST.getlist('delete_registration_docs')
                if delete_registration_ids:
                    VehicleHistoryDocument.objects.filter(
                        history=history,
                        document_type=VehicleHistoryDocument.TYPE_REGISTRATION,
                        id__in=delete_registration_ids,
                    ).delete()

                delete_insurance_casco_ids = request.POST.getlist('delete_insurance_casco_docs')
                if delete_insurance_casco_ids:
                    VehicleHistoryDocument.objects.filter(
                        history=history,
                        document_type=VehicleHistoryDocument.TYPE_INSURANCE_CASCO,
                        id__in=delete_insurance_casco_ids,
                    ).delete()

                for uploaded in request.FILES.getlist('registration_documents'):
                    if uploaded:
                        VehicleHistoryDocument.objects.create(
                            history=history,
                            document_type=VehicleHistoryDocument.TYPE_REGISTRATION,
                            file=uploaded,
                            original_filename=uploaded.name,
                        )

                for uploaded in request.FILES.getlist('insurance_casco_documents'):
                    if uploaded:
                        VehicleHistoryDocument.objects.create(
                            history=history,
                            document_type=VehicleHistoryDocument.TYPE_INSURANCE_CASCO,
                            file=uploaded,
                            original_filename=uploaded.name,
                        )

                self.message_user(request, 'Araç tarihçesi güncellendi.')
                return redirect(reverse('admin:vehicles_vehicle_history', args=[vehicle.pk]))
        else:
            form = VehicleHistoryForm(instance=history)

        context = dict(
            self.admin_site.each_context(request),
            title='Araç Tarihçesi',
            vehicle=vehicle,
            history=history,
            form=form,
            registration_docs=history.documents.filter(document_type=VehicleHistoryDocument.TYPE_REGISTRATION),
            insurance_casco_docs=history.documents.filter(document_type=VehicleHistoryDocument.TYPE_INSURANCE_CASCO),
            vehicle_notes=vehicle.notes.order_by('created_at', 'id'),
            opts=self.model._meta,
            original=vehicle,
            has_view_permission=self.has_view_permission(request, vehicle),
            has_change_permission=self.has_change_permission(request, vehicle),
            history_url=reverse('admin:vehicles_vehicle_history', args=[vehicle.pk]),
            change_url=reverse('admin:vehicles_vehicle_change', args=[vehicle.pk]),
            notes_download_url=reverse('admin:vehicles_vehicle_notes_download', args=[vehicle.pk]),
        )
        return TemplateResponse(request, 'admin/vehicles/vehicle/vehicle_history.html', context)

    def download_vehicle_notes_view(self, request, object_id):
        vehicle = self.get_object(request, object_id)
        if vehicle is None:
            return redirect('admin:vehicles_vehicle_changelist')

        notes = vehicle.notes.order_by('created_at', 'id')
        note_lines = [f'{idx}- {note.note_text}' for idx, note in enumerate(notes, start=1)]
        content = '\n'.join(note_lines) if note_lines else 'Not bulunamadı.'

        response = HttpResponse(content, content_type='text/plain; charset=utf-8')
        filename = f'{vehicle.plate}_notlar.txt'
        response['Content-Disposition'] = (
            f'attachment; filename="{filename}"; filename*=UTF-8\'\'{quote(filename)}'
        )
        return response
